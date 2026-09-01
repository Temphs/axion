"""Build the Excel workbook from the exported CSVs.

Design decisions worth knowing before editing this file:

* Every heavy statistic (medians, liquidity, scenario maths) is computed in
  Python and exported. Excel does lookups and presentation only, which is what
  keeps the workbook fast with a tape of tens of thousands of prints.
* No dynamic-array formulas. INDEX/MATCH works in every Excel, survives a
  Power Query refresh that resizes a table, and cannot silently half-spill.
* Anything a person is meant to change lives on the Settings sheet, and the
  updater reads that sheet back into config/settings.yml, so the two never
  drift apart.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from ..paths import EXPORT_DIR, WORKBOOK_FILE
from ..settings import load_settings
from . import theme as t

log = logging.getLogger(__name__)

NAV_SHEETS = [
    "Dashboard", "Holdings", "Player Terminal", "Liquidity", "Transactions",
    "Rewards", "Essence", "Investments", "Price History", "Settings", "Raw Data",
]

# Where each exported dataset lands. Visible sheets hold the tables you browse;
# hidden ones back the dashboard and the terminal.
PLACEMENTS: dict[str, tuple[str, str]] = {
    "holdings": ("Holdings", "A6"),
    "liquidity": ("Liquidity", "A6"),
    "transactions": ("Transactions", "A6"),
    "realised_trades": ("Transactions", "T6"),
    "rewards": ("Rewards", "A6"),
    "essence_summary": ("Essence", "A6"),
    "essence_ledger": ("Essence", "A14"),
    "essence_by_draw": ("Essence", "A38"),
    "investments": ("Investments", "A6"),
    "price_tape": ("Price History", "A6"),
    "kpis": ("_data_kpis", "A1"),
    "player_stats": ("_data_stats", "A1"),
    "price_tape_index": ("_data_tapeindex", "A1"),
    "positions_list": ("_data_positions", "A1"),
    "allocations": ("_data_alloc", "A1"),
    "top_exposures": ("_data_top", "A1"),
    "nav_history": ("_data_nav", "A1"),
    "rewards_by_month": ("_data_rw_month", "A1"),
    "rewards_by_competition": ("_data_rw_comp", "A1"),
    "rewards_by_scarcity": ("_data_rw_scar", "A1"),
    "meta": ("_data_meta", "A1"),
    "refresh_log": ("_data_refresh", "A1"),
    "settings": ("_data_settings", "A1"),
}

MONEY_COLUMNS = {
    "eur", "cost", "price", "value", "pl", "cash", "amount", "entry", "target", "downside", "profit",
}
PERCENT_COLUMNS = {"pct", "share", "premium", "roi", "yield"}
DATE_COLUMNS = {"at", "date", "on"}

# Enough headroom that a Power Query refresh which adds rows still lands inside
# the ranges the charts and lookups point at.
GROWTH_ROWS = 400


def _fmt_for(header: str) -> str | None:
    name = header.lower()
    if any(name.endswith(suffix) or f"_{suffix}" in name for suffix in PERCENT_COLUMNS):
        return t.PERCENT
    if any(token in name for token in MONEY_COLUMNS):
        return t.MONEY
    if name.split("_")[-1] in DATE_COLUMNS:
        return t.DATETIME
    return None


class Builder:
    def __init__(self, exports: Path = EXPORT_DIR) -> None:
        self.exports = exports
        self.settings = load_settings()
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.data: dict[str, pd.DataFrame] = {}
        self.columns: dict[str, dict[str, str]] = {}
        self.rows: dict[str, int] = {}
        self.anchor_rows: dict[str, int] = {}

    # ------------------------------------------------------------------ setup

    def load(self) -> None:
        for name in PLACEMENTS:
            path = self.exports / f"{name}.csv"
            self.data[name] = pd.read_csv(path) if path.exists() else pd.DataFrame()

    def sheet(self, name: str) -> Worksheet:
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        worksheet = self.workbook.create_sheet(name)
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.tabColor = t.ACCENT if name in NAV_SHEETS else t.BORDER
        return worksheet

    def paint(self, worksheet: Worksheet, last_row: int = 200, last_column: int = 30) -> None:
        """Dark ground. Excel has no sheet background colour, so cells carry it."""
        background = t.fill(t.BACKGROUND)
        for row in range(1, last_row + 1):
            for column in range(1, last_column + 1):
                cell = worksheet.cell(row=row, column=column)
                cell.fill = background
                cell.font = t.BODY_FONT

    # ------------------------------------------------------------------ chrome

    def header(self, worksheet: Worksheet, title: str, subtitle: str) -> None:
        worksheet["B2"] = title
        worksheet["B2"].font = t.TITLE_FONT
        worksheet["B3"] = subtitle
        worksheet["B3"].font = t.SUBTITLE_FONT
        worksheet.row_dimensions[2].height = 28

        column = 2
        for name in NAV_SHEETS:
            if name == worksheet.title:
                column += 2
                continue
            cell = worksheet.cell(row=4, column=column)
            cell.value = name
            cell.hyperlink = f"#'{name}'!A1"
            cell.font = t.font(9, colour=t.ACCENT)
            cell.fill = t.fill(t.PANEL)
            cell.alignment = t.CENTER
            cell.border = t.thin_border()
            worksheet.merge_cells(start_row=4, start_column=column, end_row=4, end_column=column + 1)
            column += 2

    def section(self, worksheet: Worksheet, row: int, text: str, note: str = "") -> None:
        cell = worksheet.cell(row=row, column=2, value=text)
        cell.font = t.SECTION_FONT
        if note:
            note_cell = worksheet.cell(row=row, column=6, value=note)
            note_cell.font = t.KPI_NOTE_FONT

    # ------------------------------------------------------------------ tables

    def write_table(self, name: str) -> None:
        sheet_name, anchor = PLACEMENTS[name]
        frame = self.data.get(name)
        worksheet = self.sheet(sheet_name)
        if frame is None or frame.empty:
            self.columns[name] = {}
            self.rows[name] = 0
            return

        letters = "".join(character for character in anchor if character.isalpha())
        anchor_row = int(anchor[len(letters):])
        anchor_column = column_index_from_string(letters)

        hidden = sheet_name.startswith("_")
        header_fill = t.fill(t.PANEL_ALT)
        for offset, column_name in enumerate(frame.columns):
            cell = worksheet.cell(row=anchor_row, column=anchor_column + offset, value=str(column_name))
            if not hidden:
                cell.font = t.HEADER_FONT
                cell.fill = header_fill
                cell.alignment = t.LEFT
                cell.border = t.thin_border()

        stripe = t.fill(t.PANEL)
        for row_offset, record in enumerate(frame.itertuples(index=False), start=1):
            for column_offset, value in enumerate(record):
                cell = worksheet.cell(
                    row=anchor_row + row_offset,
                    column=anchor_column + column_offset,
                    value=None if pd.isna(value) else value,
                )
                if hidden:
                    continue
                cell.font = t.BODY_FONT
                if row_offset % 2 == 0:
                    cell.fill = stripe
                number_format = _fmt_for(str(frame.columns[column_offset]))
                if number_format and isinstance(value, (int, float)):
                    cell.number_format = number_format

        self.columns[name] = {
            str(column): get_column_letter(anchor_column + offset)
            for offset, column in enumerate(frame.columns)
        }
        self.rows[name] = len(frame)
        self.anchor_rows[name] = anchor_row

        if not hidden:
            worksheet.freeze_panes = worksheet.cell(row=anchor_row + 1, column=1)
            for offset, column_name in enumerate(frame.columns):
                letter = get_column_letter(anchor_column + offset)
                longest = max([len(str(column_name))] + [len(str(value)) for value in frame[column_name].head(60)])
                worksheet.column_dimensions[letter].width = min(max(longest + 2, 10), 34)

    def column_range(self, dataset: str, column: str, *, absolute: bool = True) -> str:
        """A whole-column reference, so the range survives a refresh that resizes."""
        sheet_name, _ = PLACEMENTS[dataset]
        letter = self.columns.get(dataset, {}).get(column)
        if not letter:
            return '""'
        prefix = f"'{sheet_name}'!" if " " in sheet_name else f"{sheet_name}!"
        marker = "$" if absolute else ""
        return f"{prefix}{marker}{letter}:{marker}{letter}"

    def lookup(self, dataset: str, value_column: str, key_column: str, key_ref: str) -> str:
        return (
            f"IFERROR(INDEX({self.column_range(dataset, value_column)},"
            f"MATCH({key_ref},{self.column_range(dataset, key_column)},0)),\"\")"
        )

    # ------------------------------------------------------------------ charts

    @staticmethod
    def _text_properties(size: int = 900):
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import (
            CharacterProperties,
            Paragraph,
            ParagraphProperties,
        )

        properties = CharacterProperties(solidFill=t.TEXT, sz=size)
        return RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=properties), endParaRPr=properties)])

    def style_chart(self, chart, title: str, *, height: float = 7.2, width: float = 13.0):
        from openpyxl.chart.shapes import GraphicalProperties

        chart.title = title
        chart.height = height
        chart.width = width
        chart.graphical_properties = GraphicalProperties(solidFill=t.PANEL)
        chart.graphical_properties.line = LineProperties(solidFill=t.BORDER)
        chart.plot_area.graphicalProperties = GraphicalProperties(solidFill=t.PANEL)
        for axis in (chart.x_axis, chart.y_axis):
            axis.txPr = self._text_properties()
            axis.majorGridlines = None
            axis.spPr = GraphicalProperties()
            axis.spPr.line = LineProperties(solidFill=t.BORDER)
        chart.y_axis.majorGridlines = None
        if chart.legend:
            chart.legend.position = "b"
            chart.legend.txPr = self._text_properties(850)
        return chart

    def bar_chart(self, dataset: str, label_column: str, value_column: str, title: str, *, rows: int | None = None):
        sheet_name, _ = PLACEMENTS[dataset]
        worksheet = self.workbook[sheet_name]
        count = rows if rows is not None else self.rows.get(dataset, 0)
        if not count:
            return None
        anchor = self.anchor_rows.get(dataset, 1)
        label_letter = self.columns[dataset][label_column]
        value_letter = self.columns[dataset][value_column]
        chart = BarChart()
        chart.type = "bar"
        data = Reference(
            worksheet,
            min_col=column_index_from_string(value_letter),
            min_row=anchor,
            max_row=anchor + count,
        )
        categories = Reference(
            worksheet,
            min_col=column_index_from_string(label_letter),
            min_row=anchor + 1,
            max_row=anchor + count,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend = None
        from openpyxl.chart.shapes import GraphicalProperties

        chart.series[0].graphicalProperties = GraphicalProperties(solidFill=t.ACCENT)
        return self.style_chart(chart, title)

    def line_chart(self, dataset: str, label_column: str, value_columns: list[str], title: str):
        sheet_name, _ = PLACEMENTS[dataset]
        worksheet = self.workbook[sheet_name]
        count = self.rows.get(dataset, 0)
        if not count:
            return None
        anchor = self.anchor_rows.get(dataset, 1)
        chart = LineChart()
        for index, column in enumerate(value_columns):
            letter = self.columns[dataset].get(column)
            if not letter:
                continue
            data = Reference(
                worksheet,
                min_col=column_index_from_string(letter),
                min_row=anchor,
                max_row=anchor + count + GROWTH_ROWS,
            )
            chart.add_data(data, titles_from_data=True)
        label_letter = self.columns[dataset][label_column]
        chart.set_categories(
            Reference(
                worksheet,
                min_col=column_index_from_string(label_letter),
                min_row=anchor + 1,
                max_row=anchor + count + GROWTH_ROWS,
            )
        )
        palette = [t.ACCENT, t.POSITIVE, t.WARNING, t.NEGATIVE]
        from openpyxl.chart.shapes import GraphicalProperties

        for index, series in enumerate(chart.series):
            series.graphicalProperties = GraphicalProperties()
            series.graphicalProperties.line = LineProperties(solidFill=palette[index % len(palette)], w=20000)
            series.smooth = False
            series.marker = Marker(symbol="none")
        return self.style_chart(chart, title)

    # ------------------------------------------------------------------ settings

    SETTINGS_ROWS = [
        ("quick_sale_discount", "Quick-sale discount", ("valuation", "quick_sale_discount"), t.RATIO,
         "Quick-Sale Value = floor x (1 - this). 0.05 = sell 5% below floor."),
        ("fair_value_window_days", "Fair-value window (days)", ("valuation", "fair_value_window_days"), t.INTEGER,
         "1, 7, 30 or 90. Which completed-sales window sets fair value."),
        ("fair_value_inclusion_set", "Included sale types", ("valuation", "inclusion_set_name"), None,
         "SECONDARY (manager sales + accepted offers), MARKET (adds auctions and instant buys), "
         "NO_AUCTION, or ALL."),
        ("fair_value_statistic", "Statistic", ("valuation", "fair_value_statistic"), None,
         "median or mean. Median is far more robust when a single sale is an outlier."),
        ("min_sales_for_fair_value", "Minimum sales to trust", ("valuation", "min_sales_for_fair_value"), t.INTEGER,
         "Below this many prints in the window, the value is flagged LOW."),
        ("cash_balance_eur", "Cash balance (EUR)", ("account", "cash_balance_eur"), t.MONEY,
         "Sorare's API no longer exposes a stable balance field. Keep this current."),
        ("high_liquidity_min_daily_sales", "High liquidity: sales/day", ("liquidity", "high_liquidity_min_daily_sales"), t.RATIO, ""),
        ("medium_liquidity_min_daily_sales", "Medium liquidity: sales/day", ("liquidity", "medium_liquidity_min_daily_sales"), t.RATIO, ""),
        ("market_impact_warning_share", "Market-impact warning share", ("liquidity", "market_impact_warning_share"), t.RATIO,
         "Warn when your holding is this fraction of recent volume. 0.20 = 20%."),
        ("assumed_eur_per_1000_limited", "Assumed EUR / 1,000 Limited Essence", ("essence", "assumed_eur_per_1000_limited"), t.MONEY,
         "Your working assumption. The Essence sheet shows your own realised figure beside it."),
        ("assumed_eur_per_1000_rare", "Assumed EUR / 1,000 Rare Essence", ("essence", "assumed_eur_per_1000_rare"), t.MONEY, ""),
    ]

    def build_settings(self) -> None:
        from openpyxl.workbook.defined_name import DefinedName

        worksheet = self.sheet("Settings")
        self.header(
            worksheet,
            "SETTINGS",
            "Change an assumption here, save, then run update_sorare.bat. The updater reads this "
            "sheet back into config/settings.yml, so these cells are the single source of truth.",
        )
        worksheet.column_dimensions["B"].width = 34
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 96

        for column, title in (("B", "Setting"), ("C", "Value"), ("D", "What it does")):
            cell = worksheet[f"{column}6"]
            cell.value = title
            cell.font = t.HEADER_FONT
            cell.fill = t.fill(t.PANEL_ALT)
            cell.border = t.thin_border()

        settings = self.settings
        from ..settings import inclusion_set_name

        settings["valuation"]["inclusion_set_name"] = inclusion_set_name(
            settings["valuation"]["fair_value_included_types"]
        )

        for offset, (key, label, path, number_format, note) in enumerate(self.SETTINGS_ROWS):
            row = 7 + offset
            worksheet[f"B{row}"] = label
            worksheet[f"B{row}"].font = t.BODY_FONT
            value_cell = worksheet[f"C{row}"]
            value_cell.value = settings[path[0]][path[1]]
            value_cell.fill = t.fill(t.INPUT_FILL)
            value_cell.font = t.font(10, bold=True, colour=t.INPUT_TEXT)
            value_cell.border = t.thin_border(t.WARNING)
            value_cell.alignment = t.CENTER
            if number_format:
                value_cell.number_format = number_format
            worksheet[f"D{row}"] = note
            worksheet[f"D{row}"].font = t.KPI_NOTE_FONT
            worksheet[f"D{row}"].alignment = t.WRAP
            self.workbook.defined_names.add(
                DefinedName(f"SET_{key}", attr_text=f"Settings!$C${row}")
            )

        window = DataValidation(type="list", formula1='"1,7,30,90"', allow_blank=False)
        inclusion = DataValidation(type="list", formula1='"SECONDARY,MARKET,NO_AUCTION,ALL"', allow_blank=False)
        statistic = DataValidation(type="list", formula1='"median,mean"', allow_blank=False)
        for validation, cell in ((window, "C8"), (inclusion, "C9"), (statistic, "C10")):
            worksheet.add_data_validation(validation)
            validation.add(worksheet[cell])

        worksheet["B20"] = "Amber cells are the ones you edit. Everything else in the workbook follows them."
        worksheet["B20"].font = t.KPI_NOTE_FONT

    # ----------------------------------------------------------------- dashboard

    def kpi(self, metric: str) -> str:
        return (
            f"IFERROR(INDEX({self.column_range('kpis', 'value')},"
            f"MATCH(\"{metric}\",{self.column_range('kpis', 'metric')},0)),0)"
        )

    def build_dashboard(self) -> None:
        worksheet = self.sheet("Dashboard")
        self.header(
            worksheet,
            "SORARE PORTFOLIO TERMINAL",
            "Every number below is derived from your own gallery, transactions and the completed-sales "
            "tape. Values flagged LOW on Holdings are floor-derived, not traded prices.",
        )
        for column in range(2, 22):
            worksheet.column_dimensions[get_column_letter(column)].width = 13

        quick_sale = f"({self.kpi('gallery_floor_value_eur')}*(1-SET_quick_sale_discount))"
        nav = f"(SET_cash_balance_eur+{quick_sale})"
        deposits = self.kpi("deposits_eur")
        withdrawals = self.kpi("withdrawals_eur")
        economic = f"({nav}+{withdrawals}-{deposits})"
        average_capital = (
            f"IFERROR(AVERAGE({self.column_range('nav_history', 'acquisition_cost')}),"
            f"{self.kpi('acquisition_cost_eur')})"
        )

        cards = [
            ("Total deposited", f"={deposits}", t.MONEY, "manual/cash_flows.csv"),
            ("Cash balance", "=SET_cash_balance_eur", t.MONEY, "Settings sheet"),
            ("Gallery acquisition cost", f"={self.kpi('acquisition_cost_eur')}", t.MONEY, "what you paid"),
            ("Gallery market value", f"={self.kpi('market_value_eur')}", t.MONEY, "fair value x cards"),
            ("Quick-sale gallery value", f"={quick_sale}", t.MONEY, "floor less the discount"),
            ("Realized P/L", f"={self.kpi('realized_pl_eur')}", t.MONEY, "closed trades only"),
            ("Unrealized P/L", f"={self.kpi('unrealized_pl_eur')}", t.MONEY, "market value less cost"),
            ("Total rewards earned", f"={self.kpi('rewards_total_eur')}", t.MONEY, "cash + card at receipt"),
            ("Cash rewards", f"={self.kpi('cash_rewards_eur')}", t.MONEY, "actual cash"),
            ("Reward-card value", f"={self.kpi('reward_card_value_at_receipt_eur')}", t.MONEY, "valued at receipt"),
            ("Essence earned", f"={self.kpi('essence_earned')}", t.INTEGER, "Limited + Rare"),
            ("Total withdrawals", f"={withdrawals}", t.MONEY, "manual/cash_flows.csv"),
            ("Sorare NAV", f"={nav}", t.MONEY, "cash + quick-sale gallery"),
            ("True economic P/L", f"={economic}", t.MONEY, "NAV + withdrawals - deposits"),
            ("ROI", f"=IFERROR({economic}/{deposits}*100,0)", t.PERCENT, "economic P/L / deposits"),
            ("Reward yield", f"=IFERROR({self.kpi('rewards_total_eur')}/{average_capital}*100,0)", t.PERCENT,
             "rewards / average capital invested"),
        ]

        start_row = 6
        for index, (label, formula, number_format, note) in enumerate(cards):
            row = start_row + (index // 4) * 4
            column = 2 + (index % 4) * 5
            self._kpi_card(worksheet, row, column, label, formula, number_format, note)

        chart_row = start_row + 16 + 1
        self.section(worksheet, chart_row, "PORTFOLIO OVER TIME",
                     "history begins the first time the updater ran - Sorare has no portfolio history to backfill")

        charts = [
            (self.line_chart("nav_history", "taken_at", ["nav", "market_value", "quick_sale_value"],
                             "Sorare NAV over time"), f"B{chart_row + 1}"),
            (self.line_chart("nav_history", "taken_at", ["deposits", "market_value"],
                             "Cumulative deposits vs portfolio value"), f"L{chart_row + 1}"),
            (self.line_chart("rewards_by_month", "label", ["cumulative_eur"],
                             "Cumulative rewards"), f"B{chart_row + 16}"),
            (self.line_chart("nav_history", "taken_at", ["realized_pl", "unrealized_pl"],
                             "Realised vs unrealised P/L"), f"L{chart_row + 16}"),
            (self.bar_chart("allocations", "label", "value_eur", "Allocation (all dimensions)"),
             f"B{chart_row + 31}"),
            (self.bar_chart("top_exposures", "player_name", "value_eur", "Largest player exposures"),
             f"L{chart_row + 31}"),
        ]
        for chart, anchor in charts:
            if chart is not None:
                worksheet.add_chart(chart, anchor)

    def _kpi_card(
        self,
        worksheet: Worksheet,
        row: int,
        column: int,
        label: str,
        formula: str,
        number_format: str,
        note: str,
    ) -> None:
        last_column = column + 3
        worksheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=last_column)
        worksheet.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=last_column)
        worksheet.merge_cells(start_row=row + 2, start_column=column, end_row=row + 2, end_column=last_column)

        label_cell = worksheet.cell(row=row, column=column, value=label.upper())
        label_cell.font = t.KPI_LABEL_FONT
        label_cell.alignment = t.LEFT

        value_cell = worksheet.cell(row=row + 1, column=column, value=formula)
        value_cell.font = t.KPI_VALUE_FONT
        value_cell.number_format = number_format
        value_cell.alignment = t.LEFT

        note_cell = worksheet.cell(row=row + 2, column=column, value=note)
        note_cell.font = t.KPI_NOTE_FONT
        note_cell.alignment = t.LEFT

        panel = t.fill(t.PANEL)
        border = t.thin_border()
        for offset in range(3):
            for column_offset in range(column, last_column + 1):
                cell = worksheet.cell(row=row + offset, column=column_offset)
                cell.fill = panel
                cell.border = border
        worksheet.row_dimensions[row + 1].height = 24

        # Colour the P/L cards by sign, the way a terminal would.
        if "P/L" in label or label == "ROI":
            reference = value_cell.coordinate
            worksheet.conditional_formatting.add(
                reference,
                CellIsRule(operator="greaterThan", formula=["0"], font=t.font(18, bold=True, colour=t.POSITIVE)),
            )
            worksheet.conditional_formatting.add(
                reference,
                CellIsRule(operator="lessThan", formula=["0"], font=t.font(18, bold=True, colour=t.NEGATIVE)),
            )

    # ------------------------------------------------------------------ sheets

    def _conditional(self, sheet_name: str, dataset: str, column: str, rules: list) -> None:
        letter = self.columns.get(dataset, {}).get(column)
        if not letter:
            return
        worksheet = self.workbook[sheet_name]
        first_row = self.anchor_rows.get(dataset, 6) + 1
        cell_range = f"{letter}{first_row}:{letter}{first_row + self.rows.get(dataset, 0) + GROWTH_ROWS}"
        for rule in rules:
            worksheet.conditional_formatting.add(cell_range, rule)

    def build_holdings(self) -> None:
        worksheet = self.sheet("Holdings")
        self.header(
            worksheet,
            "HOLDINGS",
            "One row per Player + Scarcity + Season class. Green is profit, red is loss, amber means "
            "the valuation is thin or the player is not starting.",
        )
        green = t.font(10, colour=t.POSITIVE)
        red = t.font(10, colour=t.NEGATIVE)
        amber = t.font(10, bold=True, colour=t.WARNING)

        for column in ("unrealized_pl_eur", "unrealized_pl_pct", "return_from_cost_pct", "floor_premium_pct"):
            self._conditional(
                "Holdings", "holdings", column,
                [
                    CellIsRule(operator="greaterThan", formula=["0"], font=green),
                    CellIsRule(operator="lessThan", formula=["0"], font=red),
                ],
            )
        confidence_letter = self.columns.get("holdings", {}).get("confidence")
        if confidence_letter:
            first_row = self.anchor_rows["holdings"] + 1
            self._conditional(
                "Holdings", "holdings", "confidence",
                [FormulaRule(formula=[f'OR({confidence_letter}{first_row}="LOW",{confidence_letter}{first_row}="NONE")'],
                             font=amber, fill=t.fill(t.INPUT_FILL))],
            )
        starter_letter = self.columns.get("holdings", {}).get("starter_pct")
        if starter_letter:
            first_row = self.anchor_rows["holdings"] + 1
            self._conditional(
                "Holdings", "holdings", "starter_pct",
                [FormulaRule(formula=[f'AND({starter_letter}{first_row}<>"",{starter_letter}{first_row}<50)'],
                             font=amber)],
            )

    def build_liquidity(self) -> None:
        worksheet = self.sheet("Liquidity")
        self.header(
            worksheet,
            "LIQUIDITY",
            "Estimated Liquidation Days = cards owned / average daily completed sales. The conservative "
            "column assumes undercutting the floor only wins half the daily flow.",
        )
        self._conditional(
            "Liquidity", "liquidity", "liquidity_band",
            [
                CellIsRule(operator="equal", formula=['"HIGH"'], font=t.font(10, colour=t.POSITIVE)),
                CellIsRule(operator="equal", formula=['"MEDIUM"'], font=t.font(10, colour=t.WARNING)),
                CellIsRule(operator="equal", formula=['"LOW"'], font=t.font(10, bold=True, colour=t.NEGATIVE)),
            ],
        )
        self._conditional(
            "Liquidity", "liquidity", "market_impact_flag",
            [CellIsRule(operator="equal", formula=['"WATCH"'], font=t.font(10, bold=True, colour=t.WARNING),
                        fill=t.fill(t.INPUT_FILL))],
        )

    def build_transactions(self) -> None:
        worksheet = self.sheet("Transactions")
        self.header(
            worksheet,
            "TRANSACTIONS",
            "Every buy and sell, de-duplicated by natural key. Card-for-card trades carry "
            "is_cash_trade = 0 and are excluded from realised P/L. Closed trades are on the right.",
        )
        self._conditional(
            "Transactions", "realised_trades", "realised_pl_eur",
            [
                CellIsRule(operator="greaterThan", formula=["0"], font=t.font(10, colour=t.POSITIVE)),
                CellIsRule(operator="lessThan", formula=["0"], font=t.font(10, colour=t.NEGATIVE)),
            ],
        )
        worksheet["T5"] = "CLOSED TRADES (realised)"
        worksheet["T5"].font = t.SECTION_FONT

    def build_rewards(self) -> None:
        worksheet = self.sheet("Rewards")
        self.header(
            worksheet,
            "REWARDS",
            "Cash, reward cards and Essence are kept apart on purpose: a reward card is only worth "
            "what it sells for, so its appreciation is shown separately and never counted as cash.",
        )
        for anchor, dataset, label in (
            ("AB6", "rewards_by_competition", "By competition"),
            ("AB20", "rewards_by_scarcity", "By scarcity"),
            ("AB34", "rewards_by_month", "By month"),
        ):
            chart = self.bar_chart(dataset, "label", "total_eur", label)
            if chart is not None:
                worksheet.add_chart(chart, anchor)

    def build_essence(self) -> None:
        worksheet = self.sheet("Essence")
        self.header(
            worksheet,
            "ESSENCE",
            "Sorare's public API exposes nothing for Essence, so this reads manual/essence_log.csv. "
            "The column that matters is EUR per 1,000 - your own empirical number, Limited and Rare apart.",
        )
        self.section(worksheet, 5, "SUMMARY BY SCARCITY")
        self.section(worksheet, 13, "LEDGER")
        self.section(worksheet, 37, "BY DRAW TYPE")

    def build_investments(self) -> None:
        worksheet = self.sheet("Investments")
        self.header(
            worksheet,
            "INVESTMENTS & THESES",
            "Expected Future Price = sum of (probability x scenario price), from manual/investments.csv. "
            "A warning appears when the four probabilities do not sum to 1.",
        )
        self._conditional(
            "Investments", "investments", "expected_upside_pct",
            [
                CellIsRule(operator="greaterThan", formula=["0"], font=t.font(10, colour=t.POSITIVE)),
                CellIsRule(operator="lessThan", formula=["0"], font=t.font(10, colour=t.NEGATIVE)),
            ],
        )
        self._conditional(
            "Investments", "investments", "probability_warning",
            [CellIsRule(operator="notEqual", formula=['""'], font=t.font(10, bold=True, colour=t.WARNING))],
        )

    def build_price_history(self) -> None:
        worksheet = self.sheet("Price History")
        self.header(
            worksheet,
            "PRICE HISTORY",
            "The last 180 days of completed sales for every position you hold, sorted by position then "
            "time. The full tape stays in data/sorare.db so the workbook stays fast.",
        )

    def build_raw(self) -> None:
        worksheet = self.sheet("Raw Data")
        self.header(
            worksheet, "DATA HEALTH", "When each module last ran, and how much data is behind each sheet."
        )
        worksheet["B6"] = "Dataset"
        worksheet["C6"] = "Rows at build"
        worksheet["D6"] = "Lands on"
        for cell in ("B6", "C6", "D6"):
            worksheet[cell].font = t.HEADER_FONT
            worksheet[cell].fill = t.fill(t.PANEL_ALT)
        for offset, (name, (sheet_name, anchor)) in enumerate(sorted(PLACEMENTS.items())):
            row = 7 + offset
            worksheet[f"B{row}"] = name
            worksheet[f"C{row}"] = self.rows.get(name, 0)
            worksheet[f"D{row}"] = f"{sheet_name}!{anchor}"
            for column in "BCD":
                worksheet[f"{column}{row}"].font = t.BODY_FONT
        worksheet.column_dimensions["B"].width = 26
        worksheet.column_dimensions["C"].width = 14
        worksheet.column_dimensions["D"].width = 26

    # ------------------------------------------------------------ player terminal

    SALE_TYPES = ["MANAGER_SALE", "ACCEPTED_BUY_OFFER", "AUCTION", "INSTANT_BUY", "DIRECT_OFFER"]
    CHART_FIRST_ROW = 35
    CHART_POINTS = 150

    def build_terminal(self) -> None:
        worksheet = self.sheet("Player Terminal")
        self.header(
            worksheet,
            "PLAYER TERMINAL",
            "Pick a position, a window and which sale types count. Fair value is the median of the "
            "included completed sales - never the floor.",
        )
        for column, width in (("B", 26), ("C", 16), ("D", 16), ("E", 4), ("F", 26), ("G", 16), ("H", 16)):
            worksheet.column_dimensions[column].width = width

        positions = self.data.get("positions_list", pd.DataFrame())
        default_key = str(positions["position_key"].iloc[0]) if not positions.empty else ""

        self._terminal_controls(worksheet, default_key)
        self._terminal_metrics(worksheet)
        self._terminal_chart(worksheet)

    def _terminal_controls(self, worksheet: Worksheet, default_key: str) -> None:
        worksheet["B6"] = "POSITION"
        worksheet["B7"] = "WINDOW (DAYS)"
        worksheet["B8"] = "INCLUDED SALE TYPES"
        for row in (6, 7, 8):
            worksheet[f"B{row}"].font = t.KPI_LABEL_FONT

        worksheet["C6"] = default_key
        worksheet.merge_cells("C6:H6")
        worksheet["C7"] = self.settings["valuation"]["fair_value_window_days"]
        worksheet["C8"] = self.settings["valuation"].get("inclusion_set_name", "SECONDARY")
        for cell in ("C6", "C7", "C8"):
            worksheet[cell].fill = t.fill(t.INPUT_FILL)
            worksheet[cell].font = t.font(11, bold=True, colour=t.INPUT_TEXT)
            worksheet[cell].border = t.thin_border(t.WARNING)

        if self.rows.get("positions_list"):
            from openpyxl.workbook.defined_name import DefinedName

            # A dynamic named range rather than a whole column: the dropdown then
            # holds exactly the positions you own, with no header and no trailing
            # blanks, and it still resizes itself when a refresh changes the list.
            key_range = self.column_range("positions_list", "position_key")
            self.workbook.defined_names.add(
                DefinedName(
                    "POSITION_LIST",
                    attr_text=f"OFFSET(_data_positions!$A$2,0,0,MAX(1,COUNTA({key_range})-1),1)",
                )
            )
            validation = DataValidation(type="list", formula1="=POSITION_LIST", allow_blank=True)
            worksheet.add_data_validation(validation)
            validation.add(worksheet["C6"])
        window = DataValidation(type="list", formula1='"1,7,30,90"', allow_blank=False)
        inclusion = DataValidation(type="list", formula1='"SECONDARY,MARKET,NO_AUCTION,ALL"', allow_blank=False)
        worksheet.add_data_validation(window)
        window.add(worksheet["C7"])
        worksheet.add_data_validation(inclusion)
        inclusion.add(worksheet["C8"])

        # Helper cells, parked out of the way in column AB.
        worksheet["AB6"] = '=$C$6&" | "&$C$7&"d | "&$C$8'
        worksheet["AB7"] = (
            f'=IFERROR(MATCH($C$6,{self.column_range("price_tape_index", "position_key")},0),0)'
        )
        worksheet["AB8"] = (
            f'=IFERROR(INDEX({self.column_range("price_tape_index", "start_row")},$AB$7)'
            f'+{self.anchor_rows.get("price_tape", 6)},0)'
        )
        worksheet["AB9"] = f'=IFERROR(INDEX({self.column_range("price_tape_index", "count")},$AB$7),0)'
        for row in range(6, 10):
            worksheet[f"AA{row}"] = ["stat key", "index row", "first tape row", "prints"][row - 6]
            worksheet[f"AA{row}"].font = t.KPI_NOTE_FONT

    def _stat(self, statistic: str, window: int) -> str:
        """A number from the pre-computed grid, for the chosen inclusion set."""
        key = f'$C$6&" | {window}d | "&$C$8'
        return (
            f'=IFERROR(INDEX({self.column_range("player_stats", statistic)},'
            f'MATCH({key},{self.column_range("player_stats", "stat_key")},0)),"")'
        )

    def _holding(self, column: str) -> str:
        return f'={self.lookup("holdings", column, "position_key", "$C$6")}'

    def _terminal_metrics(self, worksheet: Worksheet) -> None:
        fair_value = self._stat("median_eur", 0).replace(" 0d ", '"&$C$7&"d ')
        # The chosen window is a cell, so the key is assembled from it directly.
        fair_value = (
            f'=IFERROR(INDEX({self.column_range("player_stats", "median_eur")},'
            f'MATCH($AB$6,{self.column_range("player_stats", "stat_key")},0)),"")'
        )
        market_rows = [
            ("Fair value (median, chosen window)", fair_value, t.MONEY),
            ("Current floor", self._holding("floor_eur"), t.MONEY),
            ("Quick-sale price", self._holding("quick_sale_price_eur"), t.MONEY),
            ("Floor premium", self._holding("floor_premium_pct"), t.PERCENT),
            ("24h average sale", self._stat("mean_eur", 1), t.MONEY),
            ("7d average sale", self._stat("mean_eur", 7), t.MONEY),
            ("7d median", self._stat("median_eur", 7), t.MONEY),
            ("30d average", self._stat("mean_eur", 30), t.MONEY),
            ("30d median", self._stat("median_eur", 30), t.MONEY),
            ("90d average", self._stat("mean_eur", 90), t.MONEY),
            ("90d median", self._stat("median_eur", 90), t.MONEY),
            ("Transactions 24h", self._stat("sales", 1), t.INTEGER),
            ("Transactions 7d", self._stat("sales", 7), t.INTEGER),
            ("Transactions 30d", self._stat("sales", 30), t.INTEGER),
            ("Valuation confidence", self._holding("confidence"), None),
        ]
        position_rows = [
            ("My cards", self._holding("cards_owned"), t.INTEGER),
            ("My average cost", self._holding("avg_cost_eur"), t.MONEY),
            ("My current value", self._holding("market_value_eur"), t.MONEY),
            ("My quick-sale value", self._holding("quick_sale_value_eur"), t.MONEY),
            ("Unrealised P/L", self._holding("unrealized_pl_eur"), t.MONEY),
            ("Return from my cost", self._holding("return_from_cost_pct"), t.PERCENT),
            ("L5", self._holding("l5"), t.RATIO),
            ("L10", self._holding("l10"), t.RATIO),
            ("L40", self._holding("l40"), t.RATIO),
            ("Starter %", self._holding("starter_pct"), t.PERCENT),
            ("Last score", self._holding("last_score"), t.RATIO),
            ("Age", self._holding("age"), t.INTEGER),
            ("Club", self._holding("club_name"), None),
            ("League", self._holding("league"), None),
            ("Position", self._holding("position"), None),
        ]

        self.section(worksheet, 10, "MARKET")
        worksheet["F10"] = "MY POSITION"
        worksheet["F10"].font = t.SECTION_FONT

        for index, (label, formula, number_format) in enumerate(market_rows):
            row = 11 + index
            self._metric_row(worksheet, row, "B", "C", label, formula, number_format)
        for index, (label, formula, number_format) in enumerate(position_rows):
            row = 11 + index
            self._metric_row(worksheet, row, "F", "G", label, formula, number_format)

        for cell in ("C14", "G16"):
            worksheet.conditional_formatting.add(
                cell, CellIsRule(operator="greaterThan", formula=["0"], font=t.font(10, colour=t.POSITIVE))
            )
            worksheet.conditional_formatting.add(
                cell, CellIsRule(operator="lessThan", formula=["0"], font=t.font(10, colour=t.NEGATIVE))
            )

    @staticmethod
    def _metric_row(
        worksheet: Worksheet,
        row: int,
        label_column: str,
        value_column: str,
        label: str,
        formula: str,
        number_format: str | None,
    ) -> None:
        label_cell = worksheet[f"{label_column}{row}"]
        label_cell.value = label
        label_cell.font = t.font(10, colour=t.TEXT_MUTED)
        value_cell = worksheet[f"{value_column}{row}"]
        value_cell.value = formula
        value_cell.font = t.font(11, bold=True)
        value_cell.alignment = t.RIGHT
        if number_format:
            value_cell.number_format = number_format
        value_cell.fill = t.fill(t.PANEL)
        value_cell.border = t.thin_border()

    def _terminal_chart(self, worksheet: Worksheet) -> None:
        self.section(
            worksheet, 32, "COMPLETED SALES",
            "each point is one completed transaction - switch a sale type off with its Y/N cell",
        )
        worksheet["B33"] = "Include? (Y/N)"
        worksheet["B33"].font = t.KPI_LABEL_FONT
        worksheet["B34"] = "Date"
        worksheet["B34"].font = t.HEADER_FONT

        toggle = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
        worksheet.add_data_validation(toggle)

        included = set(self.settings["valuation"]["fair_value_included_types"])
        for offset, sale_type in enumerate(self.SALE_TYPES):
            column = get_column_letter(3 + offset)
            worksheet[f"{column}33"] = "Y" if sale_type in included else "N"
            worksheet[f"{column}33"].fill = t.fill(t.INPUT_FILL)
            worksheet[f"{column}33"].font = t.font(10, bold=True, colour=t.INPUT_TEXT)
            worksheet[f"{column}33"].alignment = t.CENTER
            toggle.add(worksheet[f"{column}33"])
            worksheet[f"{column}34"] = sale_type
            worksheet[f"{column}34"].font = t.HEADER_FONT

        date_column = self.column_range("price_tape", "occurred_at")
        price_column = self.column_range("price_tape", "eur")
        type_column = self.column_range("price_tape", "sale_type")

        for offset in range(self.CHART_POINTS):
            row = self.CHART_FIRST_ROW + offset
            worksheet[f"B{row}"] = (
                f'=IFERROR(IF(ROW()-34>$AB$9,NA(),INDEX({date_column},$AB$8+ROW()-35)),NA())'
            )
            worksheet[f"B{row}"].number_format = t.DATE
            worksheet[f"B{row}"].font = t.font(9, colour=t.TEXT_MUTED)
            for type_offset in range(len(self.SALE_TYPES)):
                column = get_column_letter(3 + type_offset)
                worksheet[f"{column}{row}"] = (
                    f'=IFERROR(IF(AND(ROW()-34<=$AB$9,{column}$33="Y",'
                    f'INDEX({type_column},$AB$8+ROW()-35)={column}$34),'
                    f'INDEX({price_column},$AB$8+ROW()-35),NA()),NA())'
                )
                worksheet[f"{column}{row}"].number_format = t.MONEY
                worksheet[f"{column}{row}"].font = t.font(9, colour=t.TEXT_MUTED)

        from openpyxl.chart.shapes import GraphicalProperties

        chart = ScatterChart()
        chart.x_axis.number_format = t.DATE
        chart.x_axis.majorTimeUnit = "days"
        chart.style = 13
        last_row = self.CHART_FIRST_ROW + self.CHART_POINTS - 1
        x_values = Reference(worksheet, min_col=2, min_row=self.CHART_FIRST_ROW, max_row=last_row)
        palette = [t.POSITIVE, t.ACCENT, t.WARNING, "9B7BFF", t.NEGATIVE]
        for offset, sale_type in enumerate(self.SALE_TYPES):
            column_index = 3 + offset
            y_values = Reference(worksheet, min_col=column_index, min_row=34, max_row=last_row)
            series = Series(y_values, x_values, title_from_data=True)
            series.marker = Marker(symbol="circle", size=5)
            series.marker.graphicalProperties = GraphicalProperties(solidFill=palette[offset])
            series.marker.graphicalProperties.line = LineProperties(noFill=True)
            series.graphicalProperties = GraphicalProperties()
            series.graphicalProperties.line = LineProperties(noFill=True)
            chart.series.append(series)
        self.style_chart(chart, "Completed sales", height=10.5, width=26.0)
        chart.dispBlanksAs = "gap"
        worksheet.add_chart(chart, "J10")

    # ------------------------------------------------------------------- build

    # Excel has no sheet-level background colour, so the dark ground is painted
    # cell by cell. Keep these rectangles tight: every painted cell is styling
    # that has to be stored, and a generous rectangle makes the file slow to
    # open for no visual gain below the data.
    PAINT = {
        "Dashboard": (140, 26),
        "Holdings": (60, 44),
        "Player Terminal": (200, 30),
        "Liquidity": (60, 24),
        "Transactions": (60, 36),
        "Rewards": (60, 40),
        "Essence": (60, 22),
        "Investments": (60, 34),
        "Price History": (60, 8),
        "Settings": (40, 14),
        "Raw Data": (60, 12),
    }

    def build(self, path: Path = WORKBOOK_FILE) -> Path:
        self.load()

        for name in NAV_SHEETS:
            worksheet = self.sheet(name)
            rows, columns = self.PAINT[name]
            self.paint(worksheet, last_row=rows, last_column=columns)

        for name in PLACEMENTS:
            self.write_table(name)

        for sheet_name in self.workbook.sheetnames:
            if sheet_name.startswith("_"):
                self.workbook[sheet_name].sheet_state = "hidden"

        self.build_settings()
        self.build_dashboard()
        self.build_holdings()
        self.build_terminal()
        self.build_liquidity()
        self.build_transactions()
        self.build_rewards()
        self.build_essence()
        self.build_investments()
        self.build_price_history()
        self.build_raw()

        self.workbook.active = self.workbook.sheetnames.index("Dashboard")
        path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(path)
        log.info("Workbook written to %s", path)
        return path


def build_workbook(path: Path = WORKBOOK_FILE, exports: Path = EXPORT_DIR) -> Path:
    return Builder(exports).build(path)
