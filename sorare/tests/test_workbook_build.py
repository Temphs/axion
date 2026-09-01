"""The workbook build itself: it must produce a file that is actually usable.

These are the checks that catch a regression you would otherwise only find by
opening Excel - a block growing into another one, a lookup pointed at the wrong
column, a dropdown with no source.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sorare_portfolio import db, demo  # noqa: E402
from sorare_portfolio.excel import build_workbook as bw  # noqa: E402
from sorare_portfolio.export.exporter import export_all  # noqa: E402


@pytest.fixture(scope="module")
def workbook(tmp_path_factory):
    directory = tmp_path_factory.mktemp("workbook")
    exports = directory / "exports"
    exports.mkdir()

    from sorare_portfolio import paths

    original = paths.EXPORT_DIR
    import sorare_portfolio.export.exporter as exporter

    exporter.EXPORT_DIR = exports
    try:
        with db.session(directory / "demo.db") as connection:
            demo.seed(connection)
            export_all(connection)
        path = bw.Builder(exports).build(directory / "Test.xlsx")
    finally:
        exporter.EXPORT_DIR = original
    return load_workbook(path)


def test_every_expected_sheet_exists(workbook):
    for name in bw.NAV_SHEETS:
        assert name in workbook.sheetnames
    assert workbook["_data_kpis"].sheet_state == "hidden"
    assert workbook["Holdings"].sheet_state == "visible"


def test_dashboard_kpis_are_formulas_not_frozen_numbers(workbook):
    dashboard = workbook["Dashboard"]
    values = [dashboard.cell(row=row, column=column).value
              for row in (7, 11, 15, 19) for column in (2, 7, 12, 17)]
    formulas = [value for value in values if isinstance(value, str) and value.startswith("=")]
    assert len(formulas) == 16
    # NAV and economic P/L must follow the Settings sheet, not a stored constant.
    assert any("SET_quick_sale_discount" in formula for formula in formulas)
    assert any("SET_cash_balance_eur" in formula for formula in formulas)


def test_terminal_lookups_point_at_the_right_columns(workbook):
    terminal = workbook["Player Terminal"]
    assert terminal["C6"].value  # a default position is selected
    assert terminal["AB6"].value.startswith("=$C$6")
    # "My cards" must read the cards_owned column of Holdings.
    holdings = workbook["Holdings"]
    headers = {holdings.cell(row=6, column=column).value: column
               for column in range(1, holdings.max_column + 1)}
    formula = terminal["G11"].value
    expected_letter = chr(64 + headers["cards_owned"])
    assert f"Holdings!${expected_letter}$1:${expected_letter}$" in formula


def test_position_dropdown_has_a_dynamic_source(workbook):
    terminal = workbook["Player Terminal"]
    sources = [validation.formula1 for validation in terminal.data_validations.dataValidation]
    assert "=POSITION_LIST" in sources
    assert "OFFSET" in workbook.defined_names["POSITION_LIST"].attr_text


def test_layout_guard_rejects_stacked_growing_tables(tmp_path):
    builder = bw.Builder()
    builder.load()
    original = bw.PLACEMENTS["essence_by_draw"]
    bw.PLACEMENTS["essence_by_draw"] = ("Essence", "A38")
    try:
        with pytest.raises(ValueError, match="grows without bound"):
            builder.check_layout()
    finally:
        bw.PLACEMENTS["essence_by_draw"] = original


ALLOWED_FUNCTIONS = {
    "IF", "IFERROR", "AND", "OR", "NOT", "INDEX", "MATCH", "OFFSET", "COUNTA",
    "SUM", "AVERAGE", "MAX", "MIN", "ROW", "NA", "TEXT", "ROUND",
}


def _formulas(workbook):
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    yield worksheet.title, cell.coordinate, cell.value


def test_formulas_are_well_formed_and_portable(workbook):
    """Static check in place of a recalculation pass.

    Every formula must balance its brackets, reference only sheets that exist,
    and use only functions that need no _xlfn prefix - no dynamic arrays, which
    would half-spill in a file written outside Excel.
    """
    import re

    sheet_names = set(workbook.sheetnames)
    banned = {"FILTER", "SORT", "UNIQUE", "XLOOKUP", "XMATCH", "SEQUENCE", "TEXTJOIN", "LET"}
    checked = 0

    for sheet, coordinate, formula in _formulas(workbook):
        assert formula.count("(") == formula.count(")"), f"{sheet}!{coordinate}: unbalanced brackets"

        for referenced in re.findall(r"'([^']+)'!", formula) + re.findall(r"(?<![':\w])([A-Za-z_][\w]*)!", formula):
            assert referenced in sheet_names, f"{sheet}!{coordinate} references missing sheet {referenced}"

        for function in re.findall(r"([A-Z][A-Z0-9.]+)\s*\(", formula):
            assert function not in banned, f"{sheet}!{coordinate} uses {function}"
            assert function in ALLOWED_FUNCTIONS, f"{sheet}!{coordinate} uses unexpected {function}"
        checked += 1

    assert checked > 500, "expected the terminal's chart block to contribute hundreds of formulas"


def test_named_settings_resolve_to_the_settings_sheet(workbook):
    import re

    names = {name for name in workbook.defined_names if name.startswith("SET_")}
    assert len(names) == 11
    for name in names:
        assert workbook.defined_names[name].attr_text.startswith("Settings!$C$")

    used = set()
    for _, _, formula in _formulas(workbook):
        used.update(re.findall(r"SET_[a-z_0-9]+", formula))
    assert used <= names, f"formulas use undefined settings names: {used - names}"
