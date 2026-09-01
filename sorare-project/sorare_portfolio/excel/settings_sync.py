"""Read the Settings sheet back into config/settings.yml.

The Settings sheet is the interface; the YAML file is the store. Syncing on
every run means you can change an assumption in Excel, save, double-click the
updater, and see the whole workbook follow - without ever opening a text editor.

Values are rewritten line by line rather than re-serialised, so the comments in
settings.yml survive.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import load_workbook

from ..paths import SETTINGS_FILE, WORKBOOK_FILE
from ..settings import INCLUSION_SETS

log = logging.getLogger(__name__)

# Settings-sheet name -> (yaml section, yaml key)
MAPPING: dict[str, tuple[str, str]] = {
    "quick_sale_discount": ("valuation", "quick_sale_discount"),
    "fair_value_window_days": ("valuation", "fair_value_window_days"),
    "fair_value_statistic": ("valuation", "fair_value_statistic"),
    "min_sales_for_fair_value": ("valuation", "min_sales_for_fair_value"),
    "fair_value_inclusion_set": ("valuation", "fair_value_included_types"),
    "cash_balance_eur": ("account", "cash_balance_eur"),
    "high_liquidity_min_daily_sales": ("liquidity", "high_liquidity_min_daily_sales"),
    "medium_liquidity_min_daily_sales": ("liquidity", "medium_liquidity_min_daily_sales"),
    "market_impact_warning_share": ("liquidity", "market_impact_warning_share"),
    "assumed_eur_per_1000_limited": ("essence", "assumed_eur_per_1000_limited"),
    "assumed_eur_per_1000_rare": ("essence", "assumed_eur_per_1000_rare"),
}


def read_workbook_settings(path: Path = WORKBOOK_FILE) -> dict[tuple[str, str], object]:
    if not path.exists():
        return {}
    try:
        workbook = load_workbook(path, data_only=False)
    except Exception as exc:  # a workbook open in Excel, or mid-write
        log.warning("Could not read the Settings sheet (%s); keeping settings.yml as it is.", exc)
        return {}

    if "Settings" not in workbook.sheetnames:
        return {}
    worksheet = workbook["Settings"]

    values: dict[tuple[str, str], object] = {}
    for name, target in MAPPING.items():
        defined = workbook.defined_names.get(f"SET_{name}")
        if defined is None:
            continue
        for sheet_name, coordinate in defined.destinations:
            if sheet_name != "Settings":
                continue
            value = worksheet[coordinate.replace("$", "")].value
            if value is None or isinstance(value, str) and value.startswith("="):
                continue
            if name == "fair_value_inclusion_set":
                value = list(INCLUSION_SETS.get(str(value).strip().upper(), INCLUSION_SETS["SECONDARY"]))
            values[target] = value
    return values


def _replace(text: str, section: str, key: str, value: object) -> str:
    """Rewrite one `key: value` line inside one top-level section."""
    if isinstance(value, list):
        rendered = "[" + ", ".join(str(item) for item in value) + "]"
    elif isinstance(value, str):
        rendered = value
    elif isinstance(value, float) and value.is_integer() and key.endswith(("_days", "_value")):
        rendered = str(int(value))
    else:
        rendered = str(value)

    pattern = re.compile(
        rf"(^{re.escape(section)}:\n(?:[ \t]+.*\n|\n)*?[ \t]+{re.escape(key)}:[ \t]*)([^\n]*)",
        re.MULTILINE,
    )
    replaced, count = pattern.subn(lambda match: match.group(1) + rendered, text, count=1)
    if not count:
        log.warning("Setting %s.%s not found in settings.yml; leaving it alone.", section, key)
    return replaced


def sync_settings(workbook_path: Path = WORKBOOK_FILE, settings_path: Path = SETTINGS_FILE) -> int:
    values = read_workbook_settings(workbook_path)
    if not values or not settings_path.exists():
        return 0
    text = settings_path.read_text(encoding="utf-8")
    for (section, key), value in values.items():
        text = _replace(text, section, key, value)
    settings_path.write_text(text, encoding="utf-8")
    log.info("Synced %d settings from the workbook into settings.yml", len(values))
    return len(values)
