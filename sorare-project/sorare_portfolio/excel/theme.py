"""One place for every colour, font and number format in the workbook.

Dark financial-terminal palette: a near-black ground, one accent for direction,
green and red reserved exclusively for profit and loss so they always mean the
same thing, and amber only for warnings that need an action.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BACKGROUND = "0F1420"
PANEL = "171E2E"
PANEL_ALT = "1D2536"
BORDER = "2A3348"
TEXT = "E6EAF2"
TEXT_MUTED = "8A94A8"
ACCENT = "4C8DFF"
POSITIVE = "2FBF71"
NEGATIVE = "F2545B"
WARNING = "F5A623"
INPUT_FILL = "3A3116"
INPUT_TEXT = "FFD86B"

FONT_FAMILY = "Segoe UI"

MONEY = '#,##0.00 "EUR";[Red]-#,##0.00 "EUR";"-"'
MONEY_COMPACT = '#,##0 "EUR";[Red]-#,##0 "EUR";"-"'
PERCENT = '0.0"%";[Red]-0.0"%";"-"'
RATIO = '0.00'
INTEGER = '#,##0;-#,##0;"-"'
DATE = 'yyyy-mm-dd'
DATETIME = 'yyyy-mm-dd hh:mm'


def fill(colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=colour)


def font(
    size: int = 10,
    *,
    bold: bool = False,
    colour: str = TEXT,
    italic: bool = False,
) -> Font:
    return Font(name=FONT_FAMILY, size=size, bold=bold, color=colour, italic=italic)


def thin_border(colour: str = BORDER) -> Border:
    side = Side(style="thin", color=colour)
    return Border(left=side, right=side, top=side, bottom=side)


LEFT = Alignment(horizontal="left", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

TITLE_FONT = font(20, bold=True)
SUBTITLE_FONT = font(10, colour=TEXT_MUTED)
SECTION_FONT = font(12, bold=True, colour=ACCENT)
KPI_LABEL_FONT = font(9, colour=TEXT_MUTED)
KPI_VALUE_FONT = font(18, bold=True)
KPI_NOTE_FONT = font(8, colour=TEXT_MUTED, italic=True)
HEADER_FONT = font(9, bold=True, colour=TEXT)
BODY_FONT = font(10)
